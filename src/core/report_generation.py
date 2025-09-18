import json
import csv
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # Use a headless backend (no Tkinter required)
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows

from src.core.metrics import MetricsAggregator, QueryEvaluationMetrics
from src.core.configuration import SearchType


class BenchmarkReportGenerator:
    """Generate comprehensive benchmark reports in various formats"""

    def __init__(self, report_dir: Path = Path("report")):
        """
        Initialize report generator.

        Args:
            report_dir: Directory where reports will be saved (default: 'report')
        """
        self.report_dir = Path(report_dir)
        self.report_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def generate_comprehensive_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str = "benchmark_report"
    ) -> Dict[str, Path]:
        """
        Generate a comprehensive benchmark report in multiple formats.

        Args:
            metrics_aggregator: Aggregated metrics data
            report_name: Base name for report files

        Returns:
            Dictionary mapping report type to file path
        """
        report_files = {}

        # Generate different report formats
        report_files['html'] = self.generate_html_report(metrics_aggregator, report_name)
        report_files['csv'] = self.generate_csv_report(metrics_aggregator, report_name)
        report_files['json'] = self.generate_json_report(metrics_aggregator, report_name)
        report_files['markdown'] = self.generate_markdown_report(metrics_aggregator, report_name)
        report_files['excel'] = self.generate_excel_report(metrics_aggregator, report_name)

        # Generate visualizations
        try:
            report_files['plots'] = self.generate_visualization_report(metrics_aggregator, report_name)
        except Exception as e:
            print(f"Warning: Could not generate visualization report: {e}")

        return report_files

    def generate_html_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str
    ) -> Path:
        """Generate an HTML report with interactive elements"""

        html_content = self._generate_html_content(metrics_aggregator)

        # Save HTML report
        html_file = self.report_dir / f"{report_name}_{self.timestamp}.html"
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return html_file

    def generate_csv_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str
    ) -> Path:
        """Generate CSV report with detailed metrics"""

        # Convert metrics to DataFrame
        df = metrics_aggregator.to_dataframe()

        # Save detailed metrics
        csv_file = self.report_dir / f"{report_name}_detailed_{self.timestamp}.csv"
        df.to_csv(csv_file, index=False)

        # Generate summary CSV
        summary_data = []
        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        for config_id, avg_metrics in config_averages.items():
            summary_data.append({
                'configuration_id': config_id,
                'query_count': avg_metrics['query_count'],
                # Document level metrics
                'avg_document_precision': avg_metrics['document_precision'],
                'avg_document_recall': avg_metrics['document_recall'],
                'avg_document_f1_score': avg_metrics['document_f1_score'],
                'avg_document_ndcg': avg_metrics['document_ndcg'],
                'avg_document_mrr': avg_metrics['document_mrr'],
                # Section level metrics
                'avg_section_precision': avg_metrics['section_precision'],
                'avg_section_recall': avg_metrics['section_recall'],
                'avg_section_f1_score': avg_metrics['section_f1_score'],
                'avg_section_ndcg': avg_metrics['section_ndcg'],
                'avg_section_mrr': avg_metrics['section_mrr']
            })

        summary_csv_file = self.report_dir / f"{report_name}_summary_{self.timestamp}.csv"
        pd.DataFrame(summary_data).to_csv(summary_csv_file, index=False)

        return csv_file

    def generate_json_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str
    ) -> Path:
        """Generate JSON report for programmatic access"""

        # Prepare data structure
        report_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_queries': len(metrics_aggregator.metrics),
                'total_configurations': len(metrics_aggregator.get_metrics_by_configuration())
            },
            'summary': metrics_aggregator.calculate_average_metrics_by_config(),
            'detailed_metrics': [],
            'best_configurations': {}
        }

        # Add detailed metrics
        for metric in metrics_aggregator.metrics:
            report_data['detailed_metrics'].append({
                'query_id': metric.query_id,
                'query_text': metric.query_text,
                # Document level
                'document_precision': metric.document_precision,
                'document_recall': metric.document_recall,
                'document_f1_score': metric.document_f1_score,
                'document_ndcg': metric.document_ndcg,
                'document_mrr': metric.document_mrr,
                # Section level
                'section_precision': metric.section_precision,
                'section_recall': metric.section_recall,
                'section_f1_score': metric.section_f1_score,
                'section_ndcg': metric.section_ndcg,
                'section_mrr': metric.section_mrr,
                # General info
                'search_type': metric.search_type.name,
                'configuration_id': metric.configuration_id,
                'ground_truth_count': metric.ground_truth_count,
                'retrieved_count_total': metric.retrieved_count_total,
                'retrieved_count_evaluated': metric.retrieved_count_evaluated,
                'document_relevant_retrieved_count': metric.document_relevant_retrieved_count,
                'section_relevant_retrieved_count': metric.section_relevant_retrieved_count
            })

        # Add best configurations for each metric (both levels)
        for level in ['document', 'section']:
            for metric_name in ['precision', 'recall', 'f1_score', 'ndcg', 'mrr']:
                full_metric_name = f'{level}_{metric_name}'
                best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(full_metric_name)
                if best_config:
                    report_data['best_configurations'][full_metric_name] = {
                        'configuration_id': best_config,
                        'value': best_value
                    }

        # Save JSON report
        json_file = self.report_dir / f"{report_name}_{self.timestamp}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

        return json_file

    def generate_markdown_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str
    ) -> Path:
        """Generate Markdown report for documentation"""

        markdown_content = self._generate_markdown_content(metrics_aggregator)

        # Save Markdown report
        md_file = self.report_dir / f"{report_name}_{self.timestamp}.md"
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write(markdown_content)

        return md_file

    def generate_excel_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str
    ) -> Path:
        """Generate comprehensive Excel report with multiple sheets"""

        excel_file = self.report_dir / f"{report_name}_{self.timestamp}.xlsx"

        # Create workbook and remove default sheet
        wb = Workbook()
        wb.remove(wb.active)

        # Generate different sheets
        self._create_summary_sheet(wb, metrics_aggregator)
        self._create_detailed_metrics_sheet(wb, metrics_aggregator)
        self._create_configuration_comparison_sheet(wb, metrics_aggregator)
        self._create_search_type_analysis_sheet(wb, metrics_aggregator)
        self._create_query_analysis_sheet(wb, metrics_aggregator)
        self._create_performance_ranking_sheet(wb, metrics_aggregator)

        # Save workbook
        wb.save(excel_file)

        return excel_file

    def _create_summary_sheet(self, wb: Workbook, metrics_aggregator: MetricsAggregator):
        """Create executive summary sheet with the requested 5-column structure"""
        ws = wb.create_sheet("Executive Summary")

        # Title and metadata
        ws['A1'] = "RAG Benchmark Report - Executive Summary (Document vs Section Level)"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')

        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Queries: {len(metrics_aggregator.metrics)}"
        ws['A5'] = f"Total Configurations: {len(metrics_aggregator.get_metrics_by_configuration())}"
        ws[
            'A6'] = "Note: Document-level metrics (filename match) should be >= Section-level metrics (exact section match)"

        # Main comparison table with requested 5-column structure
        ws['A8'] = "Configuration Performance Comparison"
        ws['A8'].font = Font(size=14, bold=True)

        headers = [
            'Configuration',
            'Query Metrics (Document Level)',
            'Final Metrics (Document Level)',
            'Query Metrics (Section Level)',
            'Final Metrics (Section Level)'
        ]

        # Create merged headers
        ws.merge_cells('A9:A10')
        ws['A9'] = headers[0]
        ws['A9'].alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('B9:C9')
        ws['B9'] = headers[1]
        ws['B9'].alignment = Alignment(horizontal='center')

        ws.merge_cells('D9:E9')
        ws['D9'] = headers[2]
        ws['D9'].alignment = Alignment(horizontal='center')

        ws.merge_cells('F9:G9')
        ws['F9'] = headers[3]
        ws['F9'].alignment = Alignment(horizontal='center')

        ws.merge_cells('H9:I9')
        ws['H9'] = headers[4]
        ws['H9'].alignment = Alignment(horizontal='center')

        # Sub-headers for metrics
        metric_headers = ['Precision', 'Recall', 'F1-Score', 'NDCG', 'Precision', 'Recall', 'F1-Score', 'NDCG']
        for col, header in enumerate(metric_headers, 2):  # Start from column B
            ws.cell(row=10, column=col, value=header)

        # Style headers
        for row in [9, 10]:
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add configuration data
        config_averages = metrics_aggregator.calculate_average_metrics_by_config()
        for row, (config_id, metrics_data) in enumerate(config_averages.items(), 11):
            # Configuration name (shortened)
            short_config = f"Config-{row - 10}" if len(config_id) > 15 else config_id[:15]
            ws.cell(row=row, column=1, value=short_config)

            # Document level metrics (columns B-E)
            ws.cell(row=row, column=2, value=round(metrics_data['document_precision'], 4))
            ws.cell(row=row, column=3, value=round(metrics_data['document_recall'], 4))
            ws.cell(row=row, column=4, value=round(metrics_data['document_f1_score'], 4))
            ws.cell(row=row, column=5, value=round(metrics_data['document_ndcg'], 4))

            # Section level metrics (columns F-I)
            ws.cell(row=row, column=6, value=round(metrics_data['section_precision'], 4))
            ws.cell(row=row, column=7, value=round(metrics_data['section_recall'], 4))
            ws.cell(row=row, column=8, value=round(metrics_data['section_f1_score'], 4))
            ws.cell(row=row, column=9, value=round(metrics_data['section_ndcg'], 4))

            # Highlight where document level is better (should always be the case)
            for metric_col in range(2, 6):  # Document level columns
                doc_value = ws.cell(row=row, column=metric_col).value
                sec_value = ws.cell(row=row, column=metric_col + 4).value  # Corresponding section level

                if doc_value > sec_value:
                    # Green highlighting for better document performance
                    ws.cell(row=row, column=metric_col).fill = PatternFill(
                        start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"
                    )
                elif doc_value < sec_value:
                    # Red highlighting if something is wrong (shouldn't happen)
                    ws.cell(row=row, column=metric_col).fill = PatternFill(
                        start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"
                    )

        # Best performers section
        start_row = len(config_averages) + 13
        ws.cell(row=start_row, column=1, value="Best Performing Configurations").font = Font(size=14, bold=True)
        start_row += 2

        # Document level best
        ws.cell(row=start_row, column=1, value="Document Level Champions:").font = Font(bold=True)
        start_row += 1

        doc_metrics = ['document_precision', 'document_recall', 'document_f1_score', 'document_ndcg']
        for metric in doc_metrics:
            best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(metric)
            if best_config:
                metric_name = metric.replace('document_', '').replace('_', ' ').title()
                ws.cell(row=start_row, column=1, value=f"• {metric_name}:")
                ws.cell(row=start_row, column=2, value=f"{best_config[:20]}... ({best_value:.4f})")
                start_row += 1

        start_row += 1
        ws.cell(row=start_row, column=1, value="Section Level Champions:").font = Font(bold=True)
        start_row += 1

        sec_metrics = ['section_precision', 'section_recall', 'section_f1_score', 'section_ndcg']
        for metric in sec_metrics:
            best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(metric)
            if best_config:
                metric_name = metric.replace('section_', '').replace('_', ' ').title()
                ws.cell(row=start_row, column=1, value=f"• {metric_name}:")
                ws.cell(row=start_row, column=2, value=f"{best_config[:20]}... ({best_value:.4f})")
                start_row += 1

        # Auto-adjust column widths - FIXED VERSION
        column_widths = [18, 12, 10, 12, 10, 12, 10, 12, 10]  # Adjusted for the 5-column structure
        for i, width in enumerate(column_widths, 1):
            # Convert column number to letter manually to avoid MergedCell issue
            from openpyxl.utils import get_column_letter
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = width

    def _create_detailed_metrics_sheet(self, wb: Workbook, metrics_aggregator: MetricsAggregator):
        """Create detailed metrics sheet with all query results"""
        ws = wb.create_sheet("Detailed Metrics")

        # Convert to DataFrame and write to sheet
        df = metrics_aggregator.to_dataframe()

        # Add headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add conditional formatting for metrics columns
        dual_level_metrics = [
            'document_precision', 'document_recall', 'document_f1_score', 'document_ndcg', 'document_mrr',
            'section_precision', 'section_recall', 'section_f1_score', 'section_ndcg', 'section_mrr'
        ]
        for metric in dual_level_metrics:
            if metric in df.columns:
                col_idx = df.columns.get_loc(metric) + 1
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"

                rule = ColorScaleRule(
                    start_type='min', start_color='FF6B6B',  # Red for low values
                    mid_type='percentile', mid_value=50, mid_color='FFE66D',  # Yellow for medium
                    end_type='max', end_color='4ECDC4'  # Green for high values
                )
                ws.conditional_formatting.add(cell_range, rule)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_configuration_comparison_sheet(self, wb: Workbook, metrics_aggregator: MetricsAggregator):
        """Create configuration comparison sheet with charts"""
        ws = wb.create_sheet("Configuration Comparison")

        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        # Create data table for document level
        headers = ['Configuration ID', 'Doc_Precision', 'Doc_Recall', 'Doc_F1', 'Doc_NDCG', 'Sec_Precision',
                   'Sec_Recall', 'Sec_F1', 'Sec_NDCG', 'Query Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add configuration data
        config_ids = []
        for row, (config_id, metrics_data) in enumerate(config_averages.items(), 2):
            short_id = f"Config {row - 1}"
            config_ids.append(short_id)

            ws.cell(row=row, column=1, value=short_id)
            # Document level
            ws.cell(row=row, column=2, value=round(metrics_data['document_precision'], 4))
            ws.cell(row=row, column=3, value=round(metrics_data['document_recall'], 4))
            ws.cell(row=row, column=4, value=round(metrics_data['document_f1_score'], 4))
            ws.cell(row=row, column=5, value=round(metrics_data['document_ndcg'], 4))
            # Section level
            ws.cell(row=row, column=6, value=round(metrics_data['section_precision'], 4))
            ws.cell(row=row, column=7, value=round(metrics_data['section_recall'], 4))
            ws.cell(row=row, column=8, value=round(metrics_data['section_f1_score'], 4))
            ws.cell(row=row, column=9, value=round(metrics_data['section_ndcg'], 4))
            ws.cell(row=row, column=10, value=metrics_data['query_count'])

        # Create chart for document level F1 scores
        if len(config_averages) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Configuration Performance Comparison (F1 Scores)"
            chart.y_axis.title = 'F1 Score'
            chart.x_axis.title = 'Configuration'

            # Add data series for document and section F1
            doc_f1_values = Reference(ws, min_col=4, min_row=2, max_row=len(config_averages) + 1)
            sec_f1_values = Reference(ws, min_col=8, min_row=2, max_row=len(config_averages) + 1)

            doc_series = chart.add_data(doc_f1_values, titles_from_data=False)
            sec_series = chart.add_data(sec_f1_values, titles_from_data=False)

            categories = Reference(ws, min_col=1, min_row=2, max_row=len(config_averages) + 1)
            chart.set_categories(categories)
            chart.legend.position = 'b'

            ws.add_chart(chart, "L2")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_search_type_analysis_sheet(self, wb: Workbook, metrics_aggregator: MetricsAggregator):
        """Create search type analysis sheet"""
        ws = wb.create_sheet("Search Type Analysis")

        search_type_metrics = metrics_aggregator.get_metrics_by_search_type()

        # Calculate averages by search type
        search_type_averages = {}
        for search_type, metrics_list in search_type_metrics.items():
            if metrics_list:
                search_type_averages[search_type.name] = {
                    # Document level
                    'document_precision': sum(m.document_precision for m in metrics_list) / len(metrics_list),
                    'document_recall': sum(m.document_recall for m in metrics_list) / len(metrics_list),
                    'document_f1_score': sum(m.document_f1_score for m in metrics_list) / len(metrics_list),
                    'document_ndcg': sum(m.document_ndcg for m in metrics_list) / len(metrics_list),
                    # Section level
                    'section_precision': sum(m.section_precision for m in metrics_list) / len(metrics_list),
                    'section_recall': sum(m.section_recall for m in metrics_list) / len(metrics_list),
                    'section_f1_score': sum(m.section_f1_score for m in metrics_list) / len(metrics_list),
                    'section_ndcg': sum(m.section_ndcg for m in metrics_list) / len(metrics_list),
                    'query_count': len(metrics_list)
                }

        if not search_type_averages:
            ws['A1'] = "No search type data available"
            return

        # Create headers
        headers = ['Search Type', 'Doc_Precision', 'Doc_Recall', 'Doc_F1', 'Doc_NDCG',
                   'Sec_Precision', 'Sec_Recall', 'Sec_F1', 'Sec_NDCG', 'Query Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Add search type data
        for row, (search_type, metrics_data) in enumerate(search_type_averages.items(), 2):
            ws.cell(row=row, column=1, value=search_type)
            # Document level
            ws.cell(row=row, column=2, value=round(metrics_data['document_precision'], 4))
            ws.cell(row=row, column=3, value=round(metrics_data['document_recall'], 4))
            ws.cell(row=row, column=4, value=round(metrics_data['document_f1_score'], 4))
            ws.cell(row=row, column=5, value=round(metrics_data['document_ndcg'], 4))
            # Section level
            ws.cell(row=row, column=6, value=round(metrics_data['section_precision'], 4))
            ws.cell(row=row, column=7, value=round(metrics_data['section_recall'], 4))
            ws.cell(row=row, column=8, value=round(metrics_data['section_f1_score'], 4))
            ws.cell(row=row, column=9, value=round(metrics_data['section_ndcg'], 4))
            ws.cell(row=row, column=10, value=metrics_data['query_count'])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_query_analysis_sheet(self, wb: Workbook, metrics_aggregator: MetricsAggregator):
        """Create query analysis sheet showing per-query statistics"""
        ws = wb.create_sheet("Query Analysis")

        df = metrics_aggregator.to_dataframe()

        if df.empty:
            ws['A1'] = "No query data available"
            return

        # Group by query and calculate statistics
        dual_level_metrics = [
            'document_precision', 'document_recall', 'document_f1_score', 'document_ndcg',
            'section_precision', 'section_recall', 'section_f1_score', 'section_ndcg'
        ]

        query_stats = df.groupby('query_id')[dual_level_metrics].agg(['mean', 'std', 'max', 'min']).round(4)

        # 🔹 Flatten MultiIndex columns BEFORE merging
        query_stats.columns = ['_'.join(col).strip() for col in query_stats.columns.values]

        # Reset index
        query_stats = query_stats.reset_index()

        # Config counts (single-level columns already)
        config_counts = df.groupby('query_id')['configuration_id'].count().reset_index(name='config_count')

        # Safe merge now
        query_stats = query_stats.merge(config_counts, on='query_id', how='left')

        # Flatten column names for MultiIndex columns
        new_columns = []
        for col in query_stats.columns:
            if isinstance(col, tuple):
                new_columns.append('_'.join(col).strip())
            else:
                new_columns.append(col)
        query_stats.columns = new_columns

        # Write headers
        headers = list(query_stats.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Write data
        for row_idx, row_data in enumerate(query_stats.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_performance_ranking_sheet(self, wb: Workbook, metrics_aggregator: MetricsAggregator):
        """Create performance ranking sheet showing best and worst performers"""
        ws = wb.create_sheet("Performance Ranking")

        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        if not config_averages:
            ws['A1'] = "No configuration data available"
            return

        # Create rankings for each dual-level metric
        dual_level_metrics = [
            'document_precision', 'document_recall', 'document_f1_score', 'document_ndcg',
            'section_precision', 'section_recall', 'section_f1_score', 'section_ndcg'
        ]
        start_row = 1

        for metric in dual_level_metrics:
            # Sort configurations by metric
            sorted_configs = sorted(
                config_averages.items(),
                key=lambda x: x[1][metric],
                reverse=True
            )

            # Title
            ws.cell(row=start_row, column=1, value=f"{metric.replace('_', ' ').title()} Rankings").font = Font(size=14,
                                                                                                               bold=True)
            start_row += 2

            # Headers
            headers = ['Rank', 'Configuration ID', metric.replace('_', ' ').title(), 'Query Count']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            start_row += 1

            # Add ranking data
            for rank, (config_id, config_metrics) in enumerate(sorted_configs, 1):
                ws.cell(row=start_row, column=1, value=rank)
                ws.cell(row=start_row, column=2, value=config_id[:30] + "..." if len(config_id) > 30 else config_id)
                ws.cell(row=start_row, column=3, value=round(config_metrics[metric], 4))
                ws.cell(row=start_row, column=4, value=config_metrics['query_count'])

                # Highlight top 3
                if rank <= 3:
                    colors = ['FFD700', 'C0C0C0', 'CD7F32']  # Gold, Silver, Bronze
                    fill = PatternFill(start_color=colors[rank - 1], end_color=colors[rank - 1], fill_type="solid")
                    for col in range(1, 5):
                        ws.cell(row=start_row, column=col).fill = fill

                start_row += 1

            start_row += 2  # Space between sections

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    def generate_visualization_report(
            self,
            metrics_aggregator: MetricsAggregator,
            report_name: str
    ) -> Path:
        """Generate visualization plots and save as images"""

        # Create plots directory
        plots_dir = self.report_dir / f"{report_name}_plots_{self.timestamp}"
        plots_dir.mkdir(parents=True, exist_ok=True)

        # Set style
        plt.style.use('seaborn-v0_8' if 'seaborn-v0_8' in plt.style.available else 'default')

        # Generate different types of plots
        self._plot_dual_level_metrics_by_configuration(metrics_aggregator, plots_dir)
        self._plot_dual_level_metrics_by_search_type(metrics_aggregator, plots_dir)
        self._plot_dual_level_metrics_distribution(metrics_aggregator, plots_dir)
        self._plot_dual_level_configuration_comparison(metrics_aggregator, plots_dir)

        return plots_dir

    def _generate_html_content(self, metrics_aggregator: MetricsAggregator) -> str:
        """Generate HTML content for the report"""

        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>RAG Benchmark Report - Dual Level Evaluation</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }}
        .header {{ background-color: #f4f4f4; padding: 20px; border-radius: 5px; }}
        .metrics-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        .metrics-table th, .metrics-table td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        .metrics-table th {{ background-color: #f2f2f2; }}
        .best-config {{ background-color: #e7f3ff; }}
        .section {{ margin: 30px 0; }}
        .dual-level {{ background-color: #f9f9f9; }}
        .document-level {{ background-color: #e8f5e8; }}
        .section-level {{ background-color: #ffe8e8; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>RAG Benchmark Report - Dual Level Evaluation</h1>
        <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>Total Queries: {len(metrics_aggregator.metrics)}</p>
        <p>Total Configurations: {len(config_averages)}</p>
        <p><strong>Evaluation Strategy:</strong> Document-level (filename match) and Section-level (exact section match)</p>
    </div>

    <div class="section">
        <h2>Configuration Performance Summary - Document Level</h2>
        <p>Document-level metrics consider a chunk relevant if it comes from the correct filename.</p>
        <table class="metrics-table">
            <thead>
                <tr>
                    <th>Configuration ID</th>
                    <th>Queries</th>
                    <th>Precision</th>
                    <th>Recall</th>
                    <th>F1-Score</th>
                    <th>NDCG</th>
                    <th>MRR</th>
                </tr>
            </thead>
            <tbody>
        """

        # Add document-level configuration data
        for config_id, metrics in config_averages.items():
            html += f"""
                <tr class="document-level">
                    <td>{config_id[:12]}...</td>
                    <td>{metrics['query_count']}</td>
                    <td>{metrics['document_precision']:.4f}</td>
                    <td>{metrics['document_recall']:.4f}</td>
                    <td>{metrics['document_f1_score']:.4f}</td>
                    <td>{metrics['document_ndcg']:.4f}</td>
                    <td>{metrics['document_mrr']:.4f}</td>
                </tr>
            """

        html += """
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>Configuration Performance Summary - Section Level</h2>
        <p>Section-level metrics require exact chapter/section/subsection match for relevance.</p>
        <table class="metrics-table">
            <thead>
                <tr>
                    <th>Configuration ID</th>
                    <th>Queries</th>
                    <th>Precision</th>
                    <th>Recall</th>
                    <th>F1-Score</th>
                    <th>NDCG</th>
                    <th>MRR</th>
                </tr>
            </thead>
            <tbody>
        """

        # Add section-level configuration data
        for config_id, metrics in config_averages.items():
            html += f"""
                <tr class="section-level">
                    <td>{config_id[:12]}...</td>
                    <td>{metrics['query_count']}</td>
                    <td>{metrics['section_precision']:.4f}</td>
                    <td>{metrics['section_recall']:.4f}</td>
                    <td>{metrics['section_f1_score']:.4f}</td>
                    <td>{metrics['section_ndcg']:.4f}</td>
                    <td>{metrics['section_mrr']:.4f}</td>
                </tr>
            """

        html += """
            </tbody>
        </table>
    </div>

    <div class="section">
        <h2>Best Configurations by Metric</h2>
        <h3>Document Level Champions</h3>
        """

        for metric_name in ['document_precision', 'document_recall', 'document_f1_score', 'document_ndcg',
                            'document_mrr']:
            best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(metric_name)
            if best_config:
                display_name = metric_name.replace('document_', '').replace('_', ' ').title()
                html += f"<p><strong>Best {display_name}:</strong> {best_config[:12]}... ({best_value:.4f})</p>"

        html += "<h3>Section Level Champions</h3>"

        for metric_name in ['section_precision', 'section_recall', 'section_f1_score', 'section_ndcg', 'section_mrr']:
            best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(metric_name)
            if best_config:
                display_name = metric_name.replace('section_', '').replace('_', ' ').title()
                html += f"<p><strong>Best {display_name}:</strong> {best_config[:12]}... ({best_value:.4f})</p>"

        html += """
    </div>
</body>
</html>
        """

        return html

    def _generate_markdown_content(self, metrics_aggregator: MetricsAggregator) -> str:
        """Generate Markdown content for the dual-level report"""

        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        md_content = f"""# RAG Benchmark Report - Dual Level Evaluation

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  
**Total Queries:** {len(metrics_aggregator.metrics)}  
**Total Configurations:** {len(config_averages)}

## Executive Summary

This report presents the results of benchmarking different RAG configurations using dual-level evaluation:
- **Document Level:** Chunks are relevant if they come from the correct filename
- **Section Level:** Chunks are relevant only if they match the exact chapter/section/subsection

## Document Level Performance

| Configuration ID | Queries | Precision | Recall | F1-Score | NDCG | MRR |
|------------------|---------|-----------|--------|----------|------|-----|
"""

        for config_id, metrics in config_averages.items():
            md_content += f"| {config_id[:20]}... | {metrics['query_count']} | {metrics['document_precision']:.4f} | {metrics['document_recall']:.4f} | {metrics['document_f1_score']:.4f} | {metrics['document_ndcg']:.4f} | {metrics['document_mrr']:.4f} |\n"

        md_content += "\n## Section Level Performance\n\n"
        md_content += "| Configuration ID | Queries | Precision | Recall | F1-Score | NDCG | MRR |\n"
        md_content += "|------------------|---------|-----------|--------|----------|------|-----|\n"

        for config_id, metrics in config_averages.items():
            md_content += f"| {config_id[:20]}... | {metrics['query_count']} | {metrics['section_precision']:.4f} | {metrics['section_recall']:.4f} | {metrics['section_f1_score']:.4f} | {metrics['section_ndcg']:.4f} | {metrics['section_mrr']:.4f} |\n"

        md_content += "\n## Best Configurations by Metric\n\n### Document Level Champions\n\n"

        for metric_name in ['document_precision', 'document_recall', 'document_f1_score', 'document_ndcg',
                            'document_mrr']:
            best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(metric_name)
            if best_config:
                display_name = metric_name.replace('document_', '').replace('_', ' ').title()
                md_content += f"- **Best {display_name}:** {best_config} ({best_value:.4f})\n"

        md_content += "\n### Section Level Champions\n\n"

        for metric_name in ['section_precision', 'section_recall', 'section_f1_score', 'section_ndcg', 'section_mrr']:
            best_config, best_value = metrics_aggregator.get_best_configuration_by_metric(metric_name)
            if best_config:
                display_name = metric_name.replace('section_', '').replace('_', ' ').title()
                md_content += f"- **Best {display_name}:** {best_config} ({best_value:.4f})\n"

        md_content += "\n## Methodology\n\n"
        md_content += "The benchmark evaluated RAG configurations using dual-level metrics:\n\n"
        md_content += "### Document Level Metrics\n"
        md_content += "- **Precision:** Fraction of retrieved chunks from correct documents\n"
        md_content += "- **Recall:** Fraction of relevant documents that had chunks retrieved\n"
        md_content += "- **F1-Score:** Harmonic mean of document-level precision and recall\n"
        md_content += "- **NDCG:** Ranking quality considering document-level relevance\n"
        md_content += "- **MRR:** Reciprocal rank of first document-level relevant result\n\n"
        md_content += "### Section Level Metrics\n"
        md_content += "- **Precision:** Fraction of retrieved chunks from exact section matches\n"
        md_content += "- **Recall:** Fraction of relevant sections that had chunks retrieved\n"
        md_content += "- **F1-Score:** Harmonic mean of section-level precision and recall\n"
        md_content += "- **NDCG:** Ranking quality considering section-level relevance\n"
        md_content += "- **MRR:** Reciprocal rank of first section-level relevant result\n"

        return md_content

    def _plot_dual_level_metrics_by_configuration(self, metrics_aggregator: MetricsAggregator, plots_dir: Path):
        """Plot dual-level metrics comparison by configuration"""

        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        if not config_averages:
            return

        configs = list(config_averages.keys())
        config_labels = [f"Config {i + 1}" for i in range(len(configs))]

        dual_level_metrics = ['precision', 'recall', 'f1_score', 'ndcg']

        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle('Dual Level Metrics by Configuration', fontsize=16)

        for i, base_metric in enumerate(dual_level_metrics):
            ax = axes[i // 2, i % 2]

            # Get document and section level data
            doc_metric = f'document_{base_metric}'
            sec_metric = f'section_{base_metric}'

            doc_data = [config_averages[config][doc_metric] for config in configs]
            sec_data = [config_averages[config][sec_metric] for config in configs]

            x = range(len(config_labels))
            width = 0.35

            bars1 = ax.bar([i - width / 2 for i in x], doc_data, width, label='Document Level', alpha=0.8)
            bars2 = ax.bar([i + width / 2 for i in x], sec_data, width, label='Section Level', alpha=0.8)

            ax.set_title(f'{base_metric.replace("_", " ").title()}')
            ax.set_ylabel('Score')
            ax.set_xlabel('Configuration')
            ax.set_xticks(x)
            ax.set_xticklabels(config_labels, rotation=45)
            ax.legend()

            # Highlight best configurations
            best_doc_idx = doc_data.index(max(doc_data))
            best_sec_idx = sec_data.index(max(sec_data))
            bars1[best_doc_idx].set_color('orange')
            bars2[best_sec_idx].set_color('red')

        plt.tight_layout()
        plt.savefig(plots_dir / 'dual_level_metrics_by_configuration.png', dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_dual_level_metrics_by_search_type(self, metrics_aggregator: MetricsAggregator, plots_dir: Path):
        """Plot dual-level metrics comparison by search type"""

        search_type_metrics = metrics_aggregator.get_metrics_by_search_type()

        if not search_type_metrics:
            return

        # Calculate averages by search type for dual-level metrics
        search_type_averages = {}
        for search_type, metrics_list in search_type_metrics.items():
            if metrics_list:
                search_type_averages[search_type.name] = {
                    'document_precision': sum(m.document_precision for m in metrics_list) / len(metrics_list),
                    'document_recall': sum(m.document_recall for m in metrics_list) / len(metrics_list),
                    'document_f1_score': sum(m.document_f1_score for m in metrics_list) / len(metrics_list),
                    'document_ndcg': sum(m.document_ndcg for m in metrics_list) / len(metrics_list),
                    'section_precision': sum(m.section_precision for m in metrics_list) / len(metrics_list),
                    'section_recall': sum(m.section_recall for m in metrics_list) / len(metrics_list),
                    'section_f1_score': sum(m.section_f1_score for m in metrics_list) / len(metrics_list),
                    'section_ndcg': sum(m.section_ndcg for m in metrics_list) / len(metrics_list)
                }

        if not search_type_averages:
            return

        search_types = list(search_type_averages.keys())
        base_metrics = ['precision', 'recall', 'f1_score', 'ndcg']

        fig, axes = plt.subplots(2, 2, figsize=(16, 10))
        fig.suptitle('Dual Level Metrics by Search Type', fontsize=16)

        for i, base_metric in enumerate(base_metrics):
            ax = axes[i // 2, i % 2]

            doc_values = [search_type_averages[st][f'document_{base_metric}'] for st in search_types]
            sec_values = [search_type_averages[st][f'section_{base_metric}'] for st in search_types]

            x = range(len(search_types))
            width = 0.35

            ax.bar([pos - width / 2 for pos in x], doc_values, width, label='Document Level', alpha=0.8)
            ax.bar([pos + width / 2 for pos in x], sec_values, width, label='Section Level', alpha=0.8)

            ax.set_xlabel('Search Type')
            ax.set_ylabel('Score')
            ax.set_title(f'{base_metric.replace("_", " ").title()}')
            ax.set_xticks(x)
            ax.set_xticklabels(search_types)
            ax.legend()

        plt.tight_layout()
        plt.savefig(plots_dir / 'dual_level_metrics_by_search_type.png', dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_dual_level_metrics_distribution(self, metrics_aggregator: MetricsAggregator, plots_dir: Path):
        """Plot distribution of dual-level metrics values"""

        df = metrics_aggregator.to_dataframe()

        if df.empty:
            return

        fig, axes = plt.subplots(4, 2, figsize=(15, 20))
        fig.suptitle('Distribution of Dual Level Metrics', fontsize=16)

        base_metrics = ['precision', 'recall', 'f1_score', 'ndcg']

        for i, base_metric in enumerate(base_metrics):
            # Document level distribution
            ax_doc = axes[i, 0]
            doc_metric = f'document_{base_metric}'
            ax_doc.hist(df[doc_metric], bins=20, alpha=0.7, edgecolor='black', color='green')
            ax_doc.set_title(f'Document {base_metric.replace("_", " ").title()} Distribution')
            ax_doc.set_xlabel('Score')
            ax_doc.set_ylabel('Frequency')

            # Section level distribution
            ax_sec = axes[i, 1]
            sec_metric = f'section_{base_metric}'
            ax_sec.hist(df[sec_metric], bins=20, alpha=0.7, edgecolor='black', color='red')
            ax_sec.set_title(f'Section {base_metric.replace("_", " ").title()} Distribution')
            ax_sec.set_xlabel('Score')
            ax_sec.set_ylabel('Frequency')

        plt.tight_layout()
        plt.savefig(plots_dir / 'dual_level_metrics_distribution.png', dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_dual_level_configuration_comparison(self, metrics_aggregator: MetricsAggregator, plots_dir: Path):
        """Plot radar chart comparing top configurations for both levels"""

        config_averages = metrics_aggregator.calculate_average_metrics_by_config()

        if len(config_averages) < 2:
            return

        # Get top 3 configurations by document F1-score
        top_configs = sorted(config_averages.items(),
                             key=lambda x: x[1]['document_f1_score'], reverse=True)[:3]

        base_metrics = ['precision', 'recall', 'f1_score', 'ndcg']

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 10), subplot_kw=dict(projection='polar'))

        angles = [n / len(base_metrics) * 2 * 3.14159 for n in range(len(base_metrics))]
        angles += angles[:1]  # Complete the circle

        colors = ['red', 'blue', 'green']

        # Document level radar
        for i, (config_id, config_metrics) in enumerate(top_configs):
            doc_values = [config_metrics[f'document_{metric}'] for metric in base_metrics]
            doc_values += doc_values[:1]  # Complete the circle

            ax1.plot(angles, doc_values, 'o-', linewidth=2,
                     label=f'Config {i + 1}', color=colors[i])
            ax1.fill(angles, doc_values, alpha=0.25, color=colors[i])

        ax1.set_xticks(angles[:-1])
        ax1.set_xticklabels([m.replace('_', ' ').title() for m in base_metrics])
        ax1.set_ylim(0, 1)
        ax1.set_title('Top 3 Configurations - Document Level', size=16, y=1.1)
        ax1.legend(loc='upper right', bbox_to_anchor=(1.2, 1.0))

        # Section level radar
        for i, (config_id, config_metrics) in enumerate(top_configs):
            sec_values = [config_metrics[f'section_{metric}'] for metric in base_metrics]
            sec_values += sec_values[:1]  # Complete the circle

            ax2.plot(angles, sec_values, 'o-', linewidth=2,
                     label=f'Config {i + 1}', color=colors[i])
            ax2.fill(angles, sec_values, alpha=0.25, color=colors[i])

        ax2.set_xticks(angles[:-1])
        ax2.set_xticklabels([m.replace('_', ' ').title() for m in base_metrics])
        ax2.set_ylim(0, 1)
        ax2.set_title('Top 3 Configurations - Section Level', size=16, y=1.1)
        ax2.legend(loc='upper right', bbox_to_anchor=(1.2, 1.0))

        plt.tight_layout()
        plt.savefig(plots_dir / 'dual_level_configuration_comparison_radar.png', dpi=300, bbox_inches='tight')
        plt.close()